from openpyxl import load_workbook

def desensitize_name(name):
    """姓名脱敏函数"""
    length = len(name)
    if length == 2:
        return f"{name[0]}*"
    elif length == 3:
        return f"{name[0]}*{name[-1]}"
    elif length >= 4:
        star_num = length - 2
        return f"{name[0]}{'*'*star_num}{name[-1]}"
    else:
        return name

file_path = "学生名单2026级.xlsx"
wb = load_workbook(file_path)

ws1 = wb["Sheet1"]
if "Sheet2" not in wb.sheetnames:
    wb.create_sheet("Sheet2")
ws2 = wb["Sheet2"]

ws2["A1"] = "学号"
ws2["B1"] = "脱敏姓名"
#sss

row_idx = 2

for row in ws1.iter_rows(min_row=2, values_only=True):
    student_id, name = row
    # 脱敏处理
    hide_name = desensitize_name(name)
    # 写入Sheet2
    ws2[f"A{row_idx}"] = student_id
    ws2[f"B{row_idx}"] = hide_name
    row_idx += 1

wb.save(file_path)
wb.close()
print("处理完成！已生成脱敏名单到Sheet2")